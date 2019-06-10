import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def convert_aleatoire(file_number: str):
    wb = load_workbook('result/Resultats.xlsx')
    file = open("result/aleatoire/aleatoire" + file_number + ".txt")
    wb.active = 0
    ws = wb.active

    unused_cell = PatternFill(fill_type="lightGray")
    number_format = '#,##0'
    row_int = ws.max_row+1
    row_str = str(ws.max_row+1)
    temps_exe_and_oef_check = file.readline()
    while temps_exe_and_oef_check != "":
        # Taillard
        ws['A'+row_str] = int(file_number)

        # Temps d'exécution
        ws['E' + row_str] = temps_exe_and_oef_check.split(" ")[2].strip(":")

        # Plaquette ou Random
        file.readline()
        parametres = file.readline().split(":")
        if parametres[0].startswith("La"):
            research = re.search("\\[(.*)\\]", parametres[1])
            ws['B'+row_str] = "["+research.group(1)+"]"
            ws['C'+row_str].fill = unused_cell
        else:
            ws['B'+row_str].fill = unused_cell
            ws['C'+row_str] = int(parametres[1].split(",")[0].strip())

        # Nombre de pas
        ws['D' + row_str] = float(parametres[2].strip())
        ws['D' + row_str].number_format = number_format

        file.readline()
        # Fitness
        fitness_pas = file.readline().split(" ")
        ws['F' + row_str] = float(fitness_pas[2])
        ws['F' + row_str].number_format = number_format

        # Nombre de pas pour obtenir ce résultat
        ws['G' + row_str] = float(fitness_pas[4])
        ws['G' + row_str].number_format = number_format

        file.readline()
        # Solution
        '''research = re.search("\\[(.*)\\]", file.readline())
        ws['H'+row_str] = "["+research.group(1)+"]"'''

        research = re.search("\\[(.*)", file.readline())
        solution = "[" + research.group(1)
        next_line = file.readline()
        while not next_line.startswith("-----"):
            solution += next_line
            next_line = file.readline()

        ws['H' + row_str] = solution

        row_int += 1
        row_str = str(row_int)
        temps_exe_and_oef_check = file.readline()
    wb.save('result/Resultats.xlsx')
    file.close()
    wb.close()
    print(file_number, "done")


def convert_recuit(file_number: str):
    wb = load_workbook('result/Resultats.xlsx')
    file = open("result/recuit/recuit" + file_number + ".txt")
    wb.active = 1
    ws = wb.active

    unused_cell = PatternFill(fill_type="lightGray")
    number_format = '#,##0'
    row_int = ws.max_row+1
    row_str = str(ws.max_row+1)
    temps_exe_and_oef_check = file.readline()
    while temps_exe_and_oef_check != "":
        # Taillard
        ws['A'+row_str] = int(file_number)

        # Temps d'exécution
        ws['I' + row_str] = temps_exe_and_oef_check.split(" ")[2].strip(":")

        # Plaquette ou Random
        file.readline()
        parametres = file.readline().split(":")
        if parametres[0].startswith("La"):
            research = re.search("\\[(.*)\\]", parametres[1])
            ws['B'+row_str] = "["+research.group(1)+"]"
            ws['C'+row_str].fill = unused_cell
        else:
            ws['B'+row_str].fill = unused_cell
            ws['C'+row_str] = int(parametres[1].split(",")[0].strip())

        # Nb Changemt Proba
        valeurs = parametres[2].strip().split("*")
        ws['E' + row_str] = float(valeurs[0])
        ws['E' + row_str].number_format = number_format

        # Nb Chercher Voisin
        valeurs = valeurs[1].split("=")
        ws['F' + row_str] = float(valeurs[0])
        ws['F' + row_str].number_format = number_format

        # Nombre de pas
        ws['G' + row_str] = float(valeurs[1])
        ws['G' + row_str].number_format = number_format

        file.readline()
        # Fitness
        fitness_pas = file.readline().split(" ")
        ws['J' + row_str] = float(fitness_pas[2])
        ws['J' + row_str].number_format = number_format

        # Nombre de pas pour obtenir ce résultat
        ws['K' + row_str] = float(fitness_pas[4])
        ws['K' + row_str].number_format = number_format

        file.readline()
        # Solution
        '''research = re.search("\\[(.*)\\]", file.readline())
        ws['L'+row_str] = "["+research.group(1)+"]"'''

        research = re.search("\\[(.*)", file.readline())
        solution = "[" + research.group(1)
        next_line = file.readline()
        while not next_line.startswith("-----"):
            solution += next_line
            next_line = file.readline()

        ws['L' + row_str] = solution

        row_int += 1
        row_str = str(row_int)
        temps_exe_and_oef_check = file.readline()
    wb.save('result/Resultats.xlsx')
    file.close()
    wb.close()
    print(file_number, "done")


def convert_tabou(file_number: str):
    wb = load_workbook('result/Resultats.xlsx')
    file = open("result/tabou/tabou" + file_number + ".txt")
    wb.active = 2
    ws = wb.active

    unused_cell = PatternFill(fill_type="lightGray")
    number_format = '#,##0'
    row_int = ws.max_row+1
    row_str = str(ws.max_row+1)
    temps_exe_and_oef_check = file.readline()
    while temps_exe_and_oef_check != "":
        # Taillard
        ws['A'+row_str] = int(file_number)

        # Temps d'exécution
        ws['G' + row_str] = temps_exe_and_oef_check.split(" ")[2].strip(":")

        # Plaquette ou Random
        file.readline()
        parametres = file.readline().split(":")
        if parametres[0].startswith("La"):
            research = re.search("\\[(.*)\\]", parametres[1])
            ws['B'+row_str] = "["+research.group(1)+"]"
            ws['C'+row_str].fill = unused_cell
        else:
            ws['B'+row_str].fill = unused_cell
            ws['C'+row_str] = int(parametres[1].split(",")[0].strip())

        # Nombre de pas
        ws['E' + row_str] = float(parametres[2].strip())
        ws['E' + row_str].number_format = number_format

        file.readline()
        # Fitness
        fitness_pas = file.readline().split(" ")
        ws['H' + row_str] = float(fitness_pas[2])
        ws['H' + row_str].number_format = number_format

        # Nombre de pas pour obtenir ce résultat
        ws['I' + row_str] = float(fitness_pas[4])
        ws['I' + row_str].number_format = number_format

        # Nombre de fitness calculé
        nb_fitness = file.readline().split(":")
        ws['J'+row_str] = float(nb_fitness[1].strip())
        ws['J' + row_str].number_format = number_format
        # Solution
        '''research = re.search("\\[(.*)\\]", file.readline())
        ws['K'+row_str] = "["+research.group(1)+"]"'''

        research = re.search("\\[(.*)", file.readline())
        solution = "[" + research.group(1)
        next_line = file.readline()
        while not next_line.startswith("-----"):
            solution += next_line
            next_line = file.readline()

        ws['K' + row_str] = solution

        row_int += 1
        row_str = str(row_int)
        temps_exe_and_oef_check = file.readline()
    wb.save('result/Resultats.xlsx')
    file.close()
    wb.close()
    print(file_number, "done")
